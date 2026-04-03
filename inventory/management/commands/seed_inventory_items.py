import math
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from inventory.models import InventoryCategory, InventoryItem

XLSX_PATH = Path(__file__).resolve().parents[3] / "files" / "jinx-1.xlsx"


class Command(BaseCommand):
    help = "Seed InventoryCategory and InventoryItem records from files/jinx-1.xlsx"

    def handle(self, *args, **kwargs):
        try:
            wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"File not found: {XLSX_PATH}")

        if "Current Inventory" not in wb.sheetnames:  # noqa
            raise CommandError("Sheet 'Current Inventory' not found in workbook")

        ws = wb["Current Inventory"]

        created_categories = 0
        created_items = 0
        skipped_items = 0

        for row in ws.iter_rows(min_row=4, values_only=True):
            category_name = row[0]
            sku = row[1]
            item_name = row[2]
            unit_size_raw = row[3]
            reorder_point_raw = row[5]
            order_cost = row[8]
            order_count = row[9]

            # Skip rows missing required fields
            if not all([category_name, sku, item_name, unit_size_raw]):
                continue

            # Parse unit_size and uom from e.g. "384 ct", "750 mL", "1 gal"
            parts = str(unit_size_raw).split(" ", 1)
            if len(parts) != 2:
                self.stderr.write(
                    f"Skipping '{item_name}': cannot parse unit size '{unit_size_raw}'"
                )
                skipped_items += 1
                continue

            try:
                unit_size = int(parts[0])
            except ValueError:
                self.stderr.write(
                    f"Skipping '{item_name}': non-integer unit size '{parts[0]}'"
                )
                skipped_items += 1
                continue

            uom = parts[1]
            reorder_point = math.ceil(reorder_point_raw) if reorder_point_raw else 0

            category, cat_created = InventoryCategory.objects.get_or_create(
                name=category_name
            )
            if cat_created:
                created_categories += 1

            _, item_created = InventoryItem.objects.get_or_create(
                sku=sku,
                defaults={
                    "name": item_name,
                    "category": category,
                    "unit_size": unit_size,
                    "uom": uom,
                    "reorder_point": reorder_point,
                    "order_cost": order_cost,
                    "order_count": int(order_count),
                },
            )
            if item_created:
                created_items += 1
            else:
                skipped_items += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Done. Created {created_categories} categories, "
                f"{created_items} items. Skipped {skipped_items} rows."
            )
        )
