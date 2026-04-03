from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from inventory.models import InventoryItem, InventoryLog

XLSX_PATH = Path(__file__).resolve().parents[3] / "files" / "jinx-1.xlsx"


class Command(BaseCommand):
    help = "Seed InventoryLog records from files/jinx-1.xlsx"

    def handle(self, *args, **kwargs):
        try:
            wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"File not found: {XLSX_PATH}")

        if "Inventory Log" not in wb.sheetnames:
            raise CommandError("Sheet 'Inventory Log' not found in workbook")

        ws = wb["Inventory Log"]

        created_logs = 0
        skipped_rows = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            received_date = row[0]
            purchase_date = row[1]
            sku = row[2]
            quantity = row[5]
            note = row[6]

            if not all([purchase_date, sku, quantity]):
                skipped_rows += 1
                continue

            try:
                inventory_item = InventoryItem.objects.get(sku=sku)
            except InventoryItem.DoesNotExist:
                self.stderr.write(f"Skipping row: no InventoryItem with sku '{sku}'")
                skipped_rows += 1
                continue

            InventoryLog.objects.create(
                inventory_item=inventory_item,
                quantity=int(quantity),
                purchase_date=purchase_date,
                received_date=received_date,
                note=note or "",
            )
            created_logs += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Done. Created {created_logs} inventory logs. "
                f"Skipped {skipped_rows} rows."
            )
        )
